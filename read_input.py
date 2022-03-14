# Import openyxl module
import openpyxl
import datetime 
from classes import Ressource,RessourceUnavailability,Task,TaskUnavailability

INF=25

def time_process(string):
    time=string.split(":")
    if time[1][2]=='p':
        time[1]=int(time[1][:2])
        time[0]=int(time[0])
        time[0]=time[0]+12
    else:
        time[1]=int(time[1][:2])
        time[0]=int(time[0])
    time = 60*time[0]+time[1]
    return time


def read_ressources(v,ville):
    ressources={}
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("InstancesV"+str(v)+"/Instance"+ville+"V"+str(v)+".xlsx".format(ville))

    # Define variable to read the active sheet:
    worksheet = wookbook['Employees']

    # Iterate the loop to read the cell values
    for row in worksheet.iter_rows(2, max_row=INF):
        if row[0].value == None:
            break
        time1=time_process(row[5].value)
        time2=time_process(row[6].value)
        ressources[row[0].value] = Ressource(row[1].value,row[2].value,row[3].value,row[4].value,time1,time2)
    return ressources

def read_ressources_unavailabilities(v,ville,ressources):
    ressources_unavailabilities={}
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("InstancesV"+str(v)+"/Instance"+ville+"V"+str(v)+".xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook['Employees Unavailabilities']

    # Iterate the loop to read the cell values
    for row in worksheet.iter_rows(2, max_row=INF):
        if row[0].value == None:
            break
        time1=time_process(row[3].value)
        time2=time_process(row[4].value)
        ressources_unavailabilities[row[0].value] = RessourceUnavailability(row[1].value,row[2].value,time1,time2)
        ressources[row[0].value].addUnavailability(ressources_unavailabilities[row[0].value])
    return ressources_unavailabilities

def read_tasks(v,ville):
    tasks={}
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("InstancesV"+str(v)+"/Instance"+ville+"V"+str(v)+".xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook['Tasks']

    # Iterate the loop to read the cell values
    for row in worksheet.iter_rows(2, max_row=INF):
        if row[0].value == None:
            break
        time1=time_process(row[6].value)
        time2=time_process(row[7].value)
        tasks[row[0].value] = Task(row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,time1,time2)
    return tasks

def read_tasks_unavailabilities(v,ville,tasks):
    tasks_unavailabilities={}
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("InstancesV"+str(v)+"/Instance"+ville+"V"+str(v)+".xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook['Tasks Unavailabilities']

    # Iterate the loop to read the cell values
    for row in worksheet.iter_rows(2, max_row=INF):
        if row[0].value == None:
            break
        time1=time_process(row[1].value)
        time2=time_process(row[2].value)
        tasks_unavailabilities[row[0].value] = TaskUnavailability(time1,time2)
        tasks[row[0].value].addUnavailability(tasks_unavailabilities[row[0].value])
    return tasks_unavailabilities

def read_input(v,ville):
    r,t=read_ressources(v,ville),read_tasks(v,ville)
    ru,tu=read_ressources_unavailabilities(v,ville,r),read_tasks_unavailabilities(v,ville,t)
    return r,ru,t,tu